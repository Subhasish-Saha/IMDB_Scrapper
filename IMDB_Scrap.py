# Importing the required libraries
import requests
import openpyxl
from bs4 import BeautifulSoup

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'TOP_RATED_250'
print(excel.sheetnames)
sheet.append(['Rank', 'Name', 'Year of Release', 'Rating'])

try:
    source = requests.get('https://www.imdb.com/chart/top/')
    source.raise_for_status()

    soup = BeautifulSoup(source.text, 'html.parser')

    movies = soup.find('tbody', class_="lister-list").find_all('tr')

    for movie in movies:
        name = movie.find('td', class_='titleColumn').a.text
        rank = movie.find('td', class_='titleColumn').get_text(strip=True).split('.')[0]
        year = movie.find('td', class_='titleColumn').span.text.strip('()')
        # rating = movie.find('td', class_='ratingColumn imdbRating').text
        rating = movie.find('td', class_='ratingColumn imdbRating').strong.text
        sheet.append([rank, name, year, rating])

except Exception as e:
    print('We have an Error as : ', e)

excel.save('IMDB_Top_250_Movies.xlsx')
